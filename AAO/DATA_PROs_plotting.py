import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import torch

from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from DATA_PROs_functions import outcome_by_phenotype
from DATA_PROs_functions import extract_top_features

def generate_phenotype_time_evolution(rank, factor2):

    # Generate time plot for each phenotype
    timepoints = [-3,-2,-1,0,1,2,3]

    for r in range(rank):
        
        values = factor2[:, r].detach().numpy()
        plt.plot(timepoints, values[:len(timepoints)])
        plt.title(f"Time evolution of phenotype {r + 1}")
        plt.xlabel("Timepoint")
        plt.xticks(timepoints)
        plt.ylabel("Value")
        plt.grid(True)
        plt.savefig(f"./model_diagnostic_images/time_evolution_phenotype_{r + 1}.png")
        plt.clf()
        plt.close()

    return

# Real values without imputation
def plot_feature_across_phenotypes(factor0, temporal_mapping, original_T):
    
    # Features selected based on previous 0.2 contribution score cutoff
    features = ['symptom_arthritis', 'feeling_angry', 'DA_enjoyment', 'DA_leavinghome', 'feeling_worried', 
                'Lab_Alb', 'symptom_bloating', 'symptom_pain', 'Lab_Hgb', 'symptom_tired',
                'Lab_CRP', 'Lab_Wbc', 'symptom_weak', 'DA_planning', 'feeling_alone', 'DA_travel', 
                'Lab_PLT', 'feeling_frustrated']

    indices = [temporal_mapping[feature] for feature in features if feature in temporal_mapping]

    phenotypes = [15, 23, 26] # zero-indexed
    
    timepoints = [-3,-2,-1,0,1,2,3]

    means_by_phenotype = {}
    std_by_phenotype = {}  

    for p in phenotypes:
        # Get member patients for each phenotype
        column = factor0[:, p] 
        column_array = column.detach().numpy()

        pt_threshold = torch.tensor(np.percentile(column_array, 75)).item()
        member_pts = torch.nonzero(factor0[:, p] >= pt_threshold).squeeze() 
        relevant_data = original_T[member_pts]  

        means = np.nanmean(relevant_data, axis=0)  # ignore nan
        stds = np.nanstd(relevant_data, axis=0)

        means_by_phenotype[str(p)] = means
        std_by_phenotype[str(p)] = stds

    # Phenotypes selected via post-processing method
    for i, x in zip(indices, features):
        plt.figure(figsize=(10, 6))
        plt.errorbar(timepoints, means_by_phenotype['15'][i, :], yerr=std_by_phenotype['15'][i, :], fmt='-o', color='#BC272D', markersize=8, linewidth=2, capsize=5, label='Phenotype 16')
        plt.errorbar(timepoints, means_by_phenotype['23'][i, :], yerr=std_by_phenotype['23'][i, :], fmt='-s', color='#E9C716', markersize=8, linewidth=2, capsize=5, label='Phenotype 24')
        plt.errorbar(timepoints, means_by_phenotype['26'][i, :], yerr=std_by_phenotype['26'][i, :], fmt='-D', color='#50AD9F', markersize=8, linewidth=2, capsize=5, label='Phenotype 27')

        plt.xlabel('Relative Timepoint', fontsize=12)
        plt.ylabel(x, fontsize=12)
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.legend(loc='upper left', fontsize=10)
        plt.tight_layout()
        plt.savefig(f'./phenotype_feature_plots/{x}_plot.png')
        plt.close()
   
def compile_results_and_plots(temporal_phenotype_dict, static_phenotype_dict, results_df,
                              n_epoch, original_rank, rank, Lambda, lr, penalty_l1, penalty_l2,
                              lab_list, factor1, df, factor0):
    # Load existing workbook
    result_file = 'supervised_model_results.xlsx'

    with pd.ExcelWriter(result_file, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Metrics', index=False)

    wb = load_workbook(result_file)

    # Write the model hyperparamters
    hyperparameter_cols = ['n_epoch', 'model rank', 'lambda', 'lr', 'penalty_l1', 'penalty_l2']
    hyperparameters = pd.DataFrame([{
        'n_epoch': n_epoch,
        'model rank': original_rank,
        'lambda': Lambda,
        'lr': lr,
        'penalty_l1': penalty_l1,
        'penalty_l2': penalty_l2
    }])

    # Insert hyperparameters starting from row 3
    row_number = 4

    # Write the header (hyperparameter names)
    for col, column_name in enumerate(hyperparameter_cols, start=1):
        wb['Metrics'].cell(row=row_number, column=col, value=column_name)

    # Write the data (hyperparameter values)
    for col, value in enumerate(hyperparameters.iloc[0], start=1):
        wb['Metrics'].cell(row=row_number + 1, column=col, value=value)

    # Adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        value = wb['Metrics'].cell(row=1, column=['A', 'B', 'C', 'D', 'E', 'F'].index(col) + 1).value
        if value is not None:
            wb['Metrics'].column_dimensions[col].width = len(str(value)) + 4

    # Specialized workbook page for patient outcome by phenotype
    outcomes_by_phenotype = outcome_by_phenotype(wb, factor1, df, factor0)
    # Write the header
    for col_num, column_title in enumerate(outcomes_by_phenotype.columns, start=1):
        wb['Outcomes'].cell(row=1, column=col_num, value=column_title)

    # Write the data rows
    for row_num, row_data in enumerate(outcomes_by_phenotype.itertuples(index=False), start=2):
        for col_num, value in enumerate(row_data, start=1):
            wb['Outcomes'].cell(row=row_num, column=col_num, value=value)

    # Adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        value = wb['Outcomes'].cell(row=2, column=['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J'].index(col) + 1).value
        if value is not None:
            wb['Outcomes'].column_dimensions[col].width = len(str(value)) + 4

    # Process phenotypes
    for p in range(rank):
        sheet_name = f'p_{p + 1}'

        # Convert temporal and static phenotype dicts to df
        temporal_phenotype_ = temporal_phenotype_dict[f'Phenotype {p + 1}']
        temporal_phenotype_ = pd.DataFrame(temporal_phenotype_[1:], columns=temporal_phenotype_[0])

        static_phenotype_ = static_phenotype_dict[f'Phenotype {p + 1}']
        static_phenotype_ = pd.DataFrame(static_phenotype_[1:], columns=static_phenotype_[0])

        both_phenotypes = pd.concat([temporal_phenotype_, static_phenotype_], axis=1)

        # Directly load workbook and add data
        if sheet_name not in wb.sheetnames:
            phenotype_tab = wb.create_sheet(sheet_name)
        else:
            phenotype_tab = wb[sheet_name]

        for r in dataframe_to_rows(both_phenotypes, index=False, header=True):
            phenotype_tab.append(r)

        # Add temporal evolution plot
        img = Image(f'./model_diagnostic_images/time_evolution_phenotype_{p + 1}.png')
        img.width = 0.8 * img.width
        img.height = 0.8 * img.height
        phenotype_tab.add_image(img, 'E1')

        # Adjust the column widths based on content length
        for col in ['A', 'B', 'C', 'D', 'E']:
            value = phenotype_tab.cell(row=2, column=['A', 'B', 'C', 'D', 'E'].index(col) + 1).value
            if value is not None:
                phenotype_tab.column_dimensions[col].width = len(str(value)) + 3

    # Save workbook after all modifications
    wb.save(result_file)

def plot_trimmed_rf_temporal(final_phenotype_indices_outcome1, final_phenotype_indices_outcome2, factor2):
    # Union the two trimmed RF phenotype lists
    final_phenotype_indices = final_phenotype_indices_outcome1 + [item for item in final_phenotype_indices_outcome2 if item not in final_phenotype_indices_outcome1]

    timepoints = [-3,-2,-1,0,1,2,3]
    markers = ['o', 's', 'D', '^', 'v']
    muted_colors = ['lime', 'red', 'cyan', 'lime', '#4E79A7', '#8338A2', '#C13E40', '#3E7D78', '#59A14F', '#6DB6FF', '#7235B3', '#775454', '#897F21']

    filename = f"./final_phenotype_images/time_evolution_phenotypes.png"
    for p, index in zip(final_phenotype_indices, range(1,len(final_phenotype_indices) + 1)): # zero-indexed
        
        values = factor2[:, p].detach().numpy() # grab temporal evolution
        plt.plot(
            timepoints,
            values[:len(timepoints)],
            label=f"Phenotype {p + 1}",
            marker=markers[index % len(markers)],  
            color=muted_colors[index % len(muted_colors)], 
            linestyle='-', 
            markersize=6.5, 
            alpha=1 
        )

    plt.xlabel("Relative Timepoint")
    plt.xticks(timepoints)
    plt.ylabel("Value")
    plt.legend(loc='upper left', bbox_to_anchor=(1, 1), fontsize='medium')
    plt.grid(False)
    plt.tight_layout(rect=[0, 0, 0.85, 1])
    plt.savefig(filename, bbox_inches='tight')
    plt.clf()
    plt.close()