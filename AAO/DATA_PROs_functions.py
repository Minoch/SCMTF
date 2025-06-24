import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from sklearn.model_selection import train_test_split
import torch
from sklearn.cluster import SpectralClustering
from collections import Counter
import copy
from openpyxl.styles import PatternFill
import seaborn as sns
from itertools import combinations, product
from DATA_PROs_model import run_model
from DATA_PROs_classifier import RF_classifier

#### For reference ####
# factor0 samples x rank (common matrix), indicated patient membership in phenotype
# factor1 temporal features x rank
# factor2 timepoints x rank
# mfactor1 static features x rank
#######################

def get_feature_mappings(static_list, lab_list, PRO_list):
    # Get feature mappings
    static_mapping = {}
    for x in range(len(static_list)):
        static_mapping[static_list[x]] = x

    temporal_mapping = {}
    temporal_list = lab_list + PRO_list # order matters
    for y in range(len(temporal_list)):
        temporal_mapping[temporal_list[y]] = y

    return static_mapping, temporal_mapping

def plot_phenotype_importance_scores(phenotypes, temporal, important_indices):
    selected_phenotypes = {key: phenotypes[key] for idx, key in enumerate(phenotypes) if idx in important_indices}
    print(important_indices)

    for phenotype_index, (phenotype_key, phenotype_data) in enumerate(selected_phenotypes.items()):
        if temporal:
            print(f"Plotting the temporal phenotype importance scores for the one-indexed phenotype {phenotype_key}")

            title = f"Contribution Score by Temporal Features for {phenotype_key}"
            xlabel = "Temporal Feature"
            filename = f"./final_phenotype_images/temporal_features_contribution_score_{phenotype_key.replace(' ', '_').lower()}.png"
            color = '#5DADE2'
        else:
            print(f"Plotting the static phenotype importance scores for the one-indexed phenotype {phenotype_key}")

            title = f"Contribution Score by Static Features for {phenotype_key}" # one-indexed
            xlabel = "Static Feature"
            filename = f"./final_phenotype_images/static_features_contribution_score_{phenotype_key.replace(' ', '_').lower()}.png"
            color = '#8ABA96'

        phenotype_array = np.array(phenotype_data, dtype=object)
        features = phenotype_array[:, 0]  
        scores = phenotype_array[:, 1].astype(float)

        # Plotting
        plt.figure(figsize=(12, 7))
        plt.bar(features, scores, color=color)
        plt.xlabel(xlabel)
        plt.ylabel('Contribution Score')
        plt.title(title)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        plt.savefig(filename)
        plt.clf()
        plt.close()

def extract_phenotypes(phenotype_matrix, mapping, weights, temporal, threshold_fraction = 0.0):
    # Create an inverse of feature_mapping so the index is the key and the variable name is the value:
    index_to_feature_name = {value: key for key, value in mapping.items()}

    phenotypes = {}

    # Loop through each feature for the phenotype
    for col in range(phenotype_matrix.size(1)):
        column_values = phenotype_matrix[:, col]

        phenotype_data = [] # Use a list to store phenotype data

        for row_index in range(column_values.size(0)):
            value = column_values[row_index]
            dictionary_key = index_to_feature_name.get(row_index) 
            if dictionary_key:
                phenotype_data.append((dictionary_key, round(value.item(), 3)))
                # Append a tuple containing the feature name and its rounded value to the phenotype data list.
            else:
                print(f"Index {row_index} not found in dictionary")

        # Sort the phenotypes in descending order by value
        phenotype_data.sort(key=lambda item: item[1], reverse=True)

        # Store the sorted data for each phenotype
        phenotypes[f'Phenotype {col + 1}'] = phenotype_data

    return phenotypes


# 2 way phenotype overlap
def quantify_pt_overlap(factor_tensor, rank_, threshold=0.035):
    num_phenotypes = factor_tensor.shape[1] # Filter samples based on membership threshold
    overlap_matrix = np.zeros((num_phenotypes, num_phenotypes)) # Initialize overlap matrix
    
    # Get membership sets for each phenotype
    member_sets = []
    for i in range(num_phenotypes):
        member_pts = set(torch.nonzero(factor_tensor[:, i] >= threshold).squeeze().tolist())
        # print(member_pts)
        member_sets.append(member_pts)
    
    # Calculate percentage overlap
    for i in range(num_phenotypes):
        for j in range(num_phenotypes):
            intersection = member_sets[i].intersection(member_sets[j])
            min_len = min(len(member_sets[i]), len(member_sets[j]))
            overlap = len(intersection) / min_len * 100 if min_len > 0 else 0
            overlap_matrix[i, j] = overlap
    
    # Plot heatmap
    plt.figure(figsize=(10, 8))
    ax = sns.heatmap(overlap_matrix, annot=True, fmt=".1f", cmap="Blues",xticklabels=list(range(1, rank_ + 1)), yticklabels=list(range(1, rank_ + 1))) 
    plt.title("Percentage Overlap of Patients Between Phenotypes")
    plt.xlabel("Phenotype")
    plt.ylabel("Phenotype")
    plt.savefig("./model_diagnostic_images/pairwise_phenotype_overlap_heatmap.png")
    plt.show()
    return

# 3 way phenotype overlap
def quantify_triplet_overlap(factor_tensor, rank_, threshold=0.035):
    num_phenotypes = factor_tensor.shape[1]
    
    # Get membership sets for each phenotype
    member_sets = []
    for i in range(num_phenotypes):
        member_pts = set(torch.nonzero(factor_tensor[:, i] >= threshold).squeeze().tolist())
        member_sets.append(member_pts)
    
    # Calculate overlap for all triplets
    triplet_overlaps = []
    for triplet in combinations(range(num_phenotypes), 3):
        i, j, k = triplet
        intersection = member_sets[i].intersection(member_sets[j], member_sets[k])
        min_len = min(len(member_sets[i]), len(member_sets[j]), len(member_sets[k]))
        overlap = len(intersection) / min_len * 100 if min_len > 0 else 0
        triplet_overlaps.append((triplet, overlap))
    
    # Create a DataFrame and print the results as a table
    df = pd.DataFrame(triplet_overlaps, columns=["Phenotype Triplet", "Overlap (%)"])
    df = df.sort_values(by="Overlap (%)", ascending=False)  # Sort by descending order of overlap
    df.to_csv("triplet_phenotype_overlap.csv")
    print(df)

    return df

def cosine_similarity(vec_i, vec_j):
    # Calculate the cosine similarity score between 2 vectors
    cos_sim = torch.dot(vec_i, vec_j).item()
    norm_i = torch.linalg.vector_norm(vec_i, ord=2).item()
    norm_j = torch.linalg.vector_norm(vec_j, ord=2).item()
    l2norms = norm_i * norm_j 
    return cos_sim / l2norms

def normalize_phenotypes(phenotype_dict_):

    for p in phenotype_dict_:
        data = np.array(phenotype_dict_[p], dtype=object)
        values = data[:, 1].astype(float)
        
        # Compute min and max
        col_min = np.min(values)
        col_max = np.max(values)
        
        # Normalize
        if col_max != col_min:
            normalized_values = (values - col_min) / (col_max - col_min)
        else:
            normalized_values = np.zeros_like(values)  # Avoid division by zero
        
        # Update the data with normalized values
        data[:, 1] = normalized_values.astype(str)
        
        # Update the phenotype data in the dictionary with the normalized values
        phenotype_dict_[p] = [(str(data[i, 0]), float(data[i, 1])) for i in range(len(data))]
    
    return phenotype_dict_

def explore_membership_values(final_phenotype_indices_outcome1, final_phenotype_indices_outcome2, factor0):
    # Get a feel for what the patient membership values look like 
    final_phenotype_indices = final_phenotype_indices_outcome1 + [item for item in final_phenotype_indices_outcome2 if item not in final_phenotype_indices_outcome1]

    for r in final_phenotype_indices:
        column = factor0[:, r]
        column_array = column.detach().numpy()

        # Calculate the desired statistics
        mean_value = column.mean().item()
        median_value = torch.median(column).item()
        std_value = column.std().item()
        percentile_60 = torch.tensor(np.percentile(column_array, 60)).item()
        percentile_75 = torch.tensor(np.percentile(column_array, 75)).item()

        # Print the results
        print(f"Phenotype {r}")
        print(f"Mean: {mean_value}")
        print(f"Median: {median_value}")
        print(f"Standard Deviation: {std_value}")
        print(f"60th Percentile: {percentile_60}")
        print(f"75th Percentile: {percentile_75}")

    return
 
def outcome_by_phenotype(wb, phenotype_matrix, data_matrix, factor0):

    data_matrix = data_matrix.reset_index()
    
    outcomes = ['New_Med_Outcome_yr2', 'New_Med_Outcome_yr3']
    results = []
    num_phenotypes = phenotype_matrix.shape[1]

    sheet_name = 'Outcomes'
    lab_tab = wb.create_sheet(sheet_name) if sheet_name not in wb.sheetnames else wb[sheet_name]

    # Calculate percentage of positive class for each phenotype
    for phenotype_index in range(num_phenotypes):
        # Get patient membership indices for this phenotype
        column = factor0[:, phenotype_index]
        column_array = column.detach().numpy()
        percentile_75 = np.percentile(column_array, 75)
        member_pts = np.where(column_array >= percentile_75)[0]

        if len(member_pts) == 0:
            continue # If there are no members, skip this phenotype

        member_pts_list = list(member_pts)

        # Outcome1 calculations
        members1 = data_matrix.loc[member_pts_list, outcomes[0]]
        percentage_positive1 = (
            members1.sum() / len(members1) * 100
            if len(members1) > 0 else np.nan
        )

        # Outcome2 calculations
        members2 = data_matrix.loc[member_pts_list, outcomes[1]]
        percentage_positive2 = (
            members2.sum() / len(members2) * 100
            if len(members2) > 0 else np.nan
        )

        # Store result
        results.append({
            'Phenotype': phenotype_index + 1,
            '% pos. New_Med_Outcome_yr2': percentage_positive1,
            '% pos. New_Med_Outcome_yr3': percentage_positive2
        })

    results_df = pd.DataFrame(results)

    return results_df


def grid_search_models(tensor, tensor1, Mtensor, rng, device, train_idx, val_idx, test_idx, labels):

    results = []
    # Grid search
    rank_values = [10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32]
    lambda_values = [0.5, 0.6, 0.7, 0.8, 0.9]
    learning_rate_values = [0.001, 0.01, 0.1]
    penalty_l1_values = [0, 0.001, 0.01, 0.1]
    penalty_l2_values = [0]
    theta_values = [0]

    for r, lam, lr, penalty_l1, penalty_l2, theta in product(rank_values, lambda_values, learning_rate_values, penalty_l1_values, penalty_l2_values, theta_values):
        
        result_dict, hyperparameters_dict = run_model(tensor, tensor1, Mtensor, rng=rng, device=device, 
                train_idx=train_idx, val_idx=val_idx, test_idx=test_idx, labels=labels, 
                rank=r, n_epoch=1200, Lambda=lam, lr=lr, penalty_l1=penalty_l1, penalty_l2=penalty_l2, theta=theta, grid=True)
            
        # Unpack results
        MAE_val = result_dict['mae']
        RMSE_val = result_dict['rmse']
        outcome2_auc = result_dict['outcome2_auc']
        outcome3_auc = result_dict['outcome3_auc']

        # Unpack hyperparameters
        n_epoch = hyperparameters_dict['n_epoch']

        results.append((r, lam, lr, n_epoch, penalty_l1, penalty_l2, theta, MAE_val.item(), RMSE_val.item(), outcome2_auc, outcome3_auc))

    results_df = pd.DataFrame(results, columns=["Rank", "Lambda", "LearningRate", "n_epochations", 
                                "PenaltyL1", "PenaltyL2", "Theta", "MAE", "RMSE","AUC_Outcome2", "AUC_Outcome3"]) # grid search
    
    results_df.to_excel("grid_search_results.xlsx")

    # Save results to file and highlight interesting values
    writer = pd.ExcelWriter('grid_search_results_highlighted.xlsx', engine='openpyxl')
    results_df.to_excel(writer, index=False)
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']

    # Define the highlight styles
    highlight_top = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    highlight_bottom = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    def highlight_top_bottom_numbers(column_index, top=True, color_fill=highlight_top): # Function to apply highlighting
        if top:
            indices = results_df.iloc[:, column_index].nlargest(5).index # Get indices of 5 highest values
        else:
            indices = results_df.iloc[:, column_index].nsmallest(5).index # Get indices of 5 lowest values
            
        for row in indices:
            # Shift by 2 to account for the header and 0-based index
            cell = worksheet.cell(row=row+2, column=column_index+1)
            cell.fill = color_fill

    # Highlight the top 5 numbers in columns 4 and 5
    highlight_top_bottom_numbers(9, top=True, color_fill=highlight_top) # outcome 2
    highlight_top_bottom_numbers(10, top=True, color_fill=highlight_top) # outcome 3

    # Highlight the smallest 5 numbers in columns 2 and 3
    highlight_top_bottom_numbers(7, top=False, color_fill=highlight_bottom) # mae
    highlight_top_bottom_numbers(8, top=False, color_fill=highlight_bottom) # rmse

    # Save the workbook
    writer.save()
    return

def create_splits(labels2, labels3, device, seed=1221):
    # Combine labels into a single array for stratification
    combined_labels = [f"{l1}_{l2}" for l1, l2 in zip(labels2, labels3)]

    # Create indices for the full dataset
    all_indices = np.arange(len(combined_labels))

    # First split: Train and Temp (which will be split again into val and test)
    train_idx, temp_idx, _, temp_labels = train_test_split(
        all_indices,
        combined_labels,
        test_size=0.4,  # 40% goes to temp, 60% to train
        random_state=seed,
        stratify=combined_labels
    )

    # Second split: Temp into Validation and Test
    val_idx, test_idx = train_test_split(
        temp_idx,
        test_size=0.5,  # Split temp into equal parts: 20% val, 20% test
        random_state=seed,
        stratify=temp_labels
    )

    # Convert to tensor
    train_idx = torch.tensor(train_idx).to(device)
    val_idx = torch.tensor(val_idx).to(device)
    test_idx = torch.tensor(test_idx).to(device)

    return train_idx, val_idx, test_idx

def avg_phenotypes_per_pt(factor0):

    factor0_np = factor0.clone().detach().numpy()
    num_phenotypes = factor0_np.shape[1]

    for i in range(num_phenotypes):
        percentile_75_value = np.percentile(factor0_np[:, i], 75)
        factor0_np[:, i] = torch.tensor(factor0_np[:, i] >= percentile_75_value, dtype=torch.float32)

    avg_phenotypes_per_patient = factor0_np.sum(axis=1).mean().item()

    return avg_phenotypes_per_patient

def extract_top_features(phenotype_p_list, mapping, num_features=5):

    top_feature_names = [feature for feature, value in phenotype_p_list[:num_features]]
    top_feature_indices = [mapping[feature] for feature in top_feature_names if feature in mapping]

    return list(top_feature_names), top_feature_indices

def trimmed_rf_model(factor0, 
                    factor2,
                    original_data,
                    rf_shortened_phenotype_indices, 
                    temporal_phenotype_dict,
                    temporal_mapping,
                    labels, train_idx, val_idx, test_idx,
                    static_list, lab_list, PRO_list):


    final_phenotype_indices = rf_shortened_phenotype_indices  
    # print(f"final indicies {final_phenotype_indices}")    
    timepoints = [-3,-2,-1,0,1,2,3]
    feature_names = static_list+lab_list+PRO_list
    markers = ['o', 's', 'D', '^', 'v']
    muted_colors = ['#4E79A7', '#F28E2B', '#C13E40', '#3E7D78', '#59A14F', '#6DB6FF', '#7235B3', '#775454', '#897F21']
    feature_to_color = {feature: muted_colors[i % len(muted_colors)] for i, feature in enumerate(feature_names)}

    # Also save the average patient plots
    for p in final_phenotype_indices: # these are zero-indexed
        # Set pt_threshold at upper quartile
        column = factor0[:, p]
        column_array = column.detach().numpy()
        pt_threshold = torch.tensor(np.percentile(column_array, 75)).item()

        # Get the top 5 features for that phenotype
        top_5_features_dict, top_5_indices = extract_top_features(temporal_phenotype_dict[f"Phenotype {p + 1}"], temporal_mapping, num_features=5)

        filename = f"./final_phenotype_images/temporal_phenotype_{p + 1}_using_original_normalized_data.png"
        title_var = "[Original Normalized Data]"
        original_data_numpy = original_data


        means = np.nanmean(original_data_numpy, axis=0) # ignore nan
        std_devs = np.nanstd(original_data_numpy, axis=0) # ignore nan

        member_pts = torch.nonzero(factor0[:, p] >= pt_threshold).squeeze() # get the indices of member patients in this phentype
        relevant_data = original_data_numpy[member_pts] # get the desired features across all timepoints for each member
        count_non_nan = np.sum(~np.isnan(relevant_data), axis=0)
        standard_errors = std_devs / np.sqrt(count_non_nan)

        # Handle cases where there are no non-nan values for that pt for a timepoint
        with np.errstate(divide='ignore', invalid='ignore'):
            print(f"Invalid divide encountered for avg patient plotting")
            standard_errors = std_devs / np.sqrt(count_non_nan)
            standard_errors = np.where(count_non_nan == 0, 0, standard_errors)
            
        # Plotting
        plt.figure(figsize=(10, 6))
        for x in range(len(top_5_indices)):
            feature_name = top_5_features_dict[x]
            plt.errorbar(
                x=timepoints, 
                y=means[top_5_indices[x]], 
                yerr=standard_errors[top_5_indices[x]], 
                label=feature_name,
                capsize=5,
                marker=markers[x % len(markers)],  # Cycle through markers
                color=feature_to_color[feature_name],  # Cycle through muted colors
                linestyle='-',  # Line style
                markersize=6.5,  # Size of the markers
                alpha=1 # Transparency for a softer look
            )

        plt.xlabel('Relative Timepoint')
        plt.ylabel('Normalized Values') # what name?
        plt.ylim(0, 1)
        plt.legend(loc='upper right', bbox_to_anchor=(1, 1), borderaxespad=0.)
        plt.title(f'Average Normalized Patient Values for Top {len(top_5_indices)} Temporal Features for Phenotype {p + 1}') # one-indexed name
        plt.savefig(filename, bbox_inches='tight')
        plt.clf()
        plt.close()

    # Now, run RF on just those overlapping phenotypes
    _, _, importances_1, importances_2 = RF_classifier(factor0=factor0, labels=labels, train_idx=train_idx, val_idx=val_idx, test_idx=test_idx, importance_threshold=0.1, final_phenotype_indices=final_phenotype_indices)

    return final_phenotype_indices, importances_1, importances_2

